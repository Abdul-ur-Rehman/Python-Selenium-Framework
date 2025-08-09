import time
import os
from os import remove

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

#find price column and actual price of the fruit
def uploadUpdatedExcel(filePath, fruitName, colName, newPrice):
    price_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
    actual_price = driver.find_element(By.XPATH, f"//div[text()='{fruitName}']/parent::div/parent::div/div[@id='cell-{price_column}-undefined']").text
    print("Actual Price =",actual_price)


    #downloading file
    driver.find_element(By.ID, "downloadButton").click()
    time.sleep(2)


    excel = openpyxl.load_workbook(file_path)
    sheet = excel.active

    #findign price coloumn
    price_col = 0
    for col in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=col).value == colName:
            price_col = col

    # finding fruit and editing price in the excel sheet
    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row=i, column=j).value == fruitName:
                sheet.cell(row=i, column=price_col).value = newPrice

    excel.save(filePath)
    excel.close()


    #uploading file
    driver.find_element(By.CSS_SELECTOR, "input[type='file']").send_keys(filePath)
    loctor = By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)"
    WebDriverWait(driver, 5).until(EC.visibility_of_element_located(loctor))
    updated_price = driver.find_element(By.XPATH, f"//div[text()='{fruitName}']/parent::div/parent::div/div[@id='cell-{price_column}-undefined']").text


    print(driver.find_element(*loctor).text)
    print("Updated Price = ", updated_price)

driver = webdriver.Chrome()
driver.get("https://rahulshettyacademy.com/upload-download-test/")
file_path = r"C:\Users\sg\Downloads\download.xlsx"
driver.implicitly_wait(5)
uploadUpdatedExcel(file_path, "Apple", "price", 900)

if os.path.exists(file_path):
    os.remove(file_path)
    print(f"Deleted file: {file_path}")
else:
    print(f"File not found: {file_path}")

