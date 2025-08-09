import openpyxl
class HomePageTestData:

    data = [{"firstname" : "Abdur Rehman", "email" : "test@gmail.com", "password" : "Test123", "gender" : "Male", "DOB" : "09-10-1998"},
                            {"firstname" : "Jon Doe", "email" : "jon@test.com", "password" : "JONDOE123", "gender" : "Male", "DOB" : "09-10-1998"}]

    @staticmethod
    def getTestCaseData(testCase_name):
        list = []
        excel = openpyxl.load_workbook(r"D:\QA Automation\Selenium-Python\Project1_ProtoCommerce_PSF\testData\test_excel.xlsx")
        sheet = excel.active

        for i in range(1, sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == testCase_name:
                rowData = {}
                for j in range(2, sheet.max_column+1):
                    rowData[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

                list.append(rowData)

        print(list)
        return list